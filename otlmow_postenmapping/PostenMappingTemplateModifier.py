import os
import shutil
from pathlib import Path
from typing import List
from openpyxl.reader.excel import load_workbook

from otlmow_template.SubsetTemplateCreator import SubsetTemplateCreator

from otlmow_model.OtlmowModel.BaseClasses.OTLObject import dynamic_create_instance_from_uri
from otlmow_converter.DotnotationHelper import DotnotationHelper

class PostenMappingTemplateModifier(SubsetTemplateCreator):
    def __init__(self):
        super().__init__()

    @classmethod
    def alter_excel_template(cls
                             , output_path: Path
                             , instantiated_assets: List
                             , path_to_subset=None
                             , **kwargs
                             ):
        """Alter Excel template

        Alters the Excel template. Overrides the existing function and additionally removes any dummy records in Sheets.

        Args:
            output_path (Path): Excel output file that is edited
            instantiated_assets (list): List of OTLAssets
            path_to_subset (Path): Path to the Sqlite subset. Default None
        """
        generate_choice_list = kwargs.get('generate_choice_list', True)
        add_geo_artefact = kwargs.get('add_geo_artefact', False)
        add_attribute_info = kwargs.get('add_attribute_info', False)
        highlight_deprecated_attributes = kwargs.get('highlight_deprecated_attributes', False)
        amount_of_examples = kwargs.get('amount_of_examples', 0)
        delete_dummy_records = kwargs.get('delete_dummy_records', False)

        # Create a temporary output folder if not exists
        tempdir = Path(output_path).parent / 'tmpOutput'
        if not tempdir.exists():
            os.makedirs(tempdir)
        temporary_path = tempdir / 'output.xlsx'

        # Copy the output file to the temporary directory
        shutil.copyfile(output_path, temporary_path)

        wb = load_workbook(temporary_path)
        wb.create_sheet('Keuzelijsten')
        # Volgorde is belangrijk! Eerst rijen verwijderen indien nodig dan choice list toevoegen,
        # staat namelijk vast op de kolom en niet het attribuut in die kolom
        if add_geo_artefact is False:
            cls.remove_geo_artefact_excel(workbook=wb)
        if generate_choice_list:
            cls.add_choice_list_excel(workbook=wb, instantiated_attributes=instantiated_assets,
                                      path_to_subset=path_to_subset)
        if delete_dummy_records:
            cls.delete_dummy_record(workbook=wb)
        if highlight_deprecated_attributes:
            cls.check_for_deprecated_attributes(workbook=wb, instantiated_attributes=instantiated_assets)
        if add_attribute_info:
            cls.add_attribute_info_excel(workbook=wb, instantiated_attributes=instantiated_assets)
        cls.design_workbook_excel(workbook=wb)
        wb.save(output_path)
        # Remove all files from the temporary location
        file_location = os.path.dirname(temporary_path)
        [f.unlink() for f in Path(file_location).glob("*") if f.is_file()]

        # Remove the temporary folder
        os.rmdir(tempdir)

    @classmethod
    def delete_dummy_record(cls, workbook) -> None:
        """Delete the dummy records in a Workbook

        Given an Excel Workbook, delete in every sheet all dummy records.
        Dummy records have an attribute assetId.identificator starting with prefix 'dummy_'.

        Args:
            workbook (Path): The path to the Excel workbook.
        """
        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Skip the 'Keuzelijsten' sheet
            if sheet_name == "Keuzelijsten":
                continue

            # Identify the 'assetId.identificator' column
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            if "assetId.identificator" not in header_row:
                continue  # Skip if the column doesn't exist

            asset_id_col_idx = header_row.index("assetId.identificator")

            # Iterate through rows to find dummy rows to delete.
            row_indices_to_delete = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from row 2
                cell_value = row[asset_id_col_idx].value
                if cell_value and str(cell_value).startswith("dummy_"):
                    row_indices_to_delete.append(row_idx)

            # Delete rows in reverse order to avoid index shifting
            for row_idx in reversed(row_indices_to_delete):
                sheet.delete_rows(row_idx)


def create_dummy_assets(asset_list: List) -> List:
    """Create a list of dummy assets

    Given a list of OTLAssets as input, creates a List of OTLAssets, filled with dummy data.
    The output list contains one dummy record for each asset type
    """
    typeURI_set = {asset_dict.typeURI for asset_dict in asset_list}
    dummy_assets = []
    for typeURI in typeURI_set:
        instance = dynamic_create_instance_from_uri(typeURI)
        if instance is None:
            continue
        instance.fill_with_dummy_data()
        DotnotationHelper.clear_list_of_list_attributes(instance)
        dummy_assets.append(instance)
    return dummy_assets