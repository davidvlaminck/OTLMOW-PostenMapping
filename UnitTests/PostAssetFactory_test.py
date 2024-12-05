import datetime
import os.path
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

import pytest
from otlmow_model.OtlmowModel.Classes.Onderdeel.Camera import Camera
from otlmow_model.OtlmowModel.Classes.Onderdeel.RetroreflecterendeFolie import RetroreflecterendeFolie
from otlmow_model.OtlmowModel.Classes.Onderdeel.RetroreflecterendVerkeersbord import RetroreflecterendVerkeersbord
from otlmow_model.OtlmowModel.Classes.Onderdeel.WVLichtmast import WVLichtmast
from otlmow_model.OtlmowModel.BaseClasses.OTLAsset import OTLAsset
from otlmow_model.OtlmowModel.Helpers.OTLObjectHelper import is_relation
from otlmow_model.OtlmowModel.BaseClasses.FloatOrDecimalField import FloatOrDecimalField
from otlmow_model.OtlmowModel.BaseClasses.OTLField import OTLField

from otlmow_converter.OtlmowConverter import OtlmowConverter

from UnitTests.PostenMappingDict import PostenMappingDict
from UnitTests.TestModel.OtlmowModel.Classes.Onderdeel.AllCasesTestClass import AllCasesTestClass
from UnitTests.TestModelMappingDict import PostenMappingDict as TestModelPostenMappingDict
from otlmow_postenmapping.Exceptions.InvalidMappingKeyError import InvalidMappingKeyError
from otlmow_postenmapping.Exceptions.MultipleMappingKeysError import MultipleMappingKeysError
from otlmow_postenmapping.Exceptions.MissingMappingKeyError import MissingMappingKeyError

from otlmow_postenmapping.PostAssetFactory import PostAssetFactory


def set_up_factory_from_artefact() -> PostAssetFactory:
    this_file = Path(__file__)
    return PostAssetFactory(
        this_file.parent / 'Postenmapping v1.0.0 RC3 Verkeersborden-gemapt.db',
        directory=this_file.parent,
    )


def set_up_factory_with_unittest_mapping() -> PostAssetFactory:
    factory = PostAssetFactory()
    factory.mapping_dict = PostenMappingDict.mapping_dict
    return factory


def test_load_factory_with_default_mapping():
    factory = PostAssetFactory()
    assert factory.mapping_dict is not None


def test_load_vkb_postenmapping():
    factory = set_up_factory_with_unittest_mapping()
    mapping = factory.mapping_dict
    assert '1001.10714' in mapping
    assert '1001.20111' in mapping

    assert mapping['1001.10714']['None'][
               'typeURI'] == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'
    assert mapping['1001.20111']['2'][
               'typeURI'] == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'

    assert mapping['1001.10714'] == {
        'None':
             {'attributen': {'diameter':
                                 {'dotnotation': 'diameter',
                                  'range': None,
                                  'type': 'http://www.w3.org/2001/XMLSchema#decimal',
                                  'typeURI': 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun.diameter',
                                  'value': '114'},
                             'lengte':
                                 {'dotnotation': 'lengte',
                                  'range': None,
                                  'type': 'http://www.w3.org/2001/XMLSchema#decimal',
                                  'typeURI': 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun.lengte',
                                  'value': None}},
              'isHoofdAsset': False,
              'typeURI': 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'}}


def test_create_assets_from_post_1001_20111(subtests):
    factory = set_up_factory_with_unittest_mapping()
    created_assets = factory.create_assets_from_post('1001.20111')

    folie_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendeFolie'
                      for a in created_assets)
    bord_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'
                     for a in created_assets)
    steun_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'
                      for a in created_assets)

    assert folie_count == 1
    assert bord_count == 1
    assert steun_count == 1

    bord = next((a for a in created_assets if a.typeURI ==
                 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'), None)

    assert bord is not None

    with subtests.test(msg='correct value (x <= 0.5), not raising ValueError'):
        bord.oppervlakte.waarde = 0.5
        assert True

    with subtests.test(msg='incorrect value, raising ValueError'):
        with pytest.raises(ValueError):
            bord.oppervlakte.waarde = 0.6


def test_create_assets_from_post_1001_20128(subtests):
    factory = set_up_factory_with_unittest_mapping()
    created_assets = factory.create_assets_from_post('1001.20128')

    folie_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendeFolie'
                      for a in created_assets)
    bord_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'
                     for a in created_assets)
    steun_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'
                      for a in created_assets)

    assert folie_count == 1
    assert bord_count == 1
    assert steun_count == 0

    bord = next((a for a in created_assets if a.typeURI ==
                 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'), None)

    assert bord is not None

    with subtests.test(msg='correct value (0.5 < x <= 1), not raising ValueError'):
        bord.oppervlakte.waarde = 0.6
        assert True

    with subtests.test(msg='incorrect value, raising ValueError'):
        with pytest.raises(ValueError):
            bord.oppervlakte.waarde = 0.2
        with pytest.raises(ValueError):
            bord.oppervlakte.waarde = 1.6


def test_create_assets_from_post_1001_20131(subtests):
    factory = set_up_factory_with_unittest_mapping()
    created_assets = factory.create_assets_from_post('1001.20131')

    folie_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendeFolie'
                      for a in created_assets)
    bord_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'
                     for a in created_assets)
    steun_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'
                      for a in created_assets)

    assert folie_count == 1
    assert bord_count == 1
    assert steun_count == 1

    folie = next((a for a in created_assets if a.typeURI ==
                  'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendeFolie'), None)

    assert folie is not None

    with subtests.test(msg='correct value, not raising ValueError'):
        folie.folietype = 'folietype-3a'
        assert True

    with subtests.test(msg='incorrect value, raising ValueError'):
        with pytest.raises(ValueError):
            folie.folietype = 'folietype-1'

def test_get_valid_template_key_from_base_asset_happy_flow(subtests):
    with subtests.test(msg='1 bestekPostNummer beschikbaar'):
        factory = set_up_factory_with_unittest_mapping() # maakt zelf een factory die specifiek gemaakt is voor uw test
        asset = Camera() # maakt een asset aan die specifiek gemaakt is voor uw test
        asset.bestekPostNummer = ['1001.10111']

        template_key = factory.get_valid_mapping_key_from_base_asset(asset)

        assert template_key == '1001.10111'

    with subtests.test(msg='Meerdere bestekPostNummers beschikbaar, waarvan minstens 1 geldig'):
        factory = set_up_factory_with_unittest_mapping() # maakt zelf een factory die specifiek gemaakt is voor uw test
        asset = Camera() # maakt een asset aan die specifiek gemaakt is voor uw test
        asset.bestekPostNummer = ['1001.10111', 'randomBestekPostNummer']

        template_key = factory.get_valid_mapping_key_from_base_asset(asset)

        assert template_key == '1001.10111'


def test_get_valid_template_key_from_base_asset_invalid_template_key():
    # arrange opzetten test scenario
    factory = set_up_factory_with_unittest_mapping()  # maakt zelf een factory die specifiek gemaakt is voor uw test
    asset = Camera()  # maakt een asset aan die specifiek gemaakt is voor uw test
    asset.bestekPostNummer = ['invalid_template_key']

    # act + assert
    with pytest.raises(InvalidMappingKeyError):
        factory.get_valid_mapping_key_from_base_asset(asset)

def test_get_valid_template_key_from_base_asset_multiple_template_keys():
    # arrange opzetten test scenario
    factory = set_up_factory_with_unittest_mapping()  # maakt zelf een factory die specifiek gemaakt is voor uw test
    asset = Camera()  # maakt een asset aan die specifiek gemaakt is voor uw test
    asset.bestekPostNummer = ['1001.10111', '1001.10112']

    # act + assert
    with pytest.raises(MultipleMappingKeysError):
        factory.get_valid_mapping_key_from_base_asset(asset)

def test_get_valid_template_key_from_base_asset_missing_template_key():
    # arrange opzetten test scenario
    factory = set_up_factory_with_unittest_mapping()  # maakt zelf een factory die specifiek gemaakt is voor uw test
    asset = Camera()  # maakt een asset aan die specifiek gemaakt is voor uw test
    asset.bestekPostNummer = None

    # act + assert
    with pytest.raises(MissingMappingKeyError):
        factory.get_valid_mapping_key_from_base_asset(asset)

def test_create_assets_from_post_1001_30704(subtests):
    factory = set_up_factory_with_unittest_mapping()
    created_assets = factory.create_assets_from_post('1001.30704')

    folie_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendeFolie'
                      for a in created_assets)
    bord_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'
                     for a in created_assets)
    steun_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'
                      for a in created_assets)

    assert folie_count == 0
    assert bord_count == 0
    assert steun_count == 1

    steun = next((a for a in created_assets if a.typeURI ==
                  'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'), None)

    assert steun is not None

    with subtests.test(msg='correct value for float/decimal'):
        assert steun.diameter.waarde == 114.0


def test_create_assets_using_testclass():
    model_directory_path = Path(__file__).parent / 'TestModel'
    factory = PostAssetFactory()
    factory.mapping_dict = TestModelPostenMappingDict.mapping_dict

    test_class_base = AllCasesTestClass()
    test_class_base.bestekPostNummer = ['testclass_1']

    created_assets = factory.create_assets_from_mapping(test_class_base, unique_index=0,
                                                          overwrite_original_attributes_by_mapping=True,
                                                        model_directory=model_directory_path)

    testclass = next((a for a in created_assets if a.typeURI ==
                      'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#AllCasesTestClass'), None)

    assert testclass is not None

    # Simple datatypes (eenvoudige datatypes)
    assert testclass.testBooleanField == True
    assert testclass.testIntegerField == 9
    assert testclass.testStringField == 'myDummyString'
    assert testclass.testDateField == date(2000, 1, 1)
    assert testclass.testDateTimeField == datetime.datetime(year=2000, month=1, day=1, hour=1, minute=1, second=1)
    assert testclass.testKwantWrd.waarde == 1.1
    assert testclass.testKeuzelijst == 'waarde-1'
    assert testclass.testIntegerFieldMetKard[0] == 1
    assert testclass.testStringFieldMetKard[0] == 'myDummyString1'
    # Complex datatypes (complexe datatypes)
    assert testclass.testComplexType.testBooleanField == True
    assert testclass.testComplexType.testStringField == 'myDummyString'
    # Union datatypes (union datatypes)
    assert testclass.testUnionType.unionString == 'myDummyString'
    assert testclass.testUnionType.unionKwantWrd.waarde == 1.1

def test_create_assets_from_post_1001_10171(subtests):
    factory = set_up_factory_with_unittest_mapping()
    created_assets = factory.create_assets_from_post('1001.10171')

    folie_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendeFolie'
                      for a in created_assets)
    bord_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'
                     for a in created_assets)
    steun_count = sum(a.typeURI == 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#Verkeersbordsteun'
                      for a in created_assets)

    assert folie_count == 1
    assert bord_count == 1
    assert steun_count == 0

    bord = next((a for a in created_assets if a.typeURI ==
                 'https://wegenenverkeer.data.vlaanderen.be/ns/onderdeel#RetroreflecterendVerkeersbord'), None)

    assert bord is not None

    with subtests.test(msg='correct value for uniontype'):
        assert bord.afmeting.achthoekig.zijde.waarde == 400


def test_split_range_str(subtests):
    with subtests.test(msg='x < 0.5'):
        expected_list = [('st', 0.5)]
        assert PostAssetFactory._split_numeric_range_str('x < 0.5') == expected_list

    with subtests.test(msg='x <= 0.5'):
        expected_list = [('ste', 0.5)]
        assert PostAssetFactory._split_numeric_range_str('x <= 0.5') == expected_list

    with subtests.test(msg='0.5 < x'):
        expected_list = [('gt', 0.5)]
        assert PostAssetFactory._split_numeric_range_str('0.5 < x') == expected_list

    with subtests.test(msg='0.5 <= x'):
        expected_list = [('gte', 0.5)]
        assert PostAssetFactory._split_numeric_range_str('0.5 <= x') == expected_list

    with subtests.test(msg='0.5 < x <= 1'):
        expected_list = [('gt', 0.5), ('ste', 1)]
        assert PostAssetFactory._split_numeric_range_str('0.5 < x <= 1') == expected_list


def test_create_asset_from_mapping_happy_flow(subtests):
    this_directory = Path(__file__).parent
    factory = PostAssetFactory(this_directory / 'PostenMapping_template_202411.db', directory=this_directory, mapping_name='PostenMapping_template_202411')
    my_wvlichtmast = WVLichtmast()
    my_wvlichtmast.fill_with_dummy_data()
    my_wvlichtmast.naam = 'myDummyNaam'
    my_wvlichtmast.bestekPostNummer = ['WVlichtmast_config1']

    my_list_OTLObjects = factory.create_assets_from_mapping(my_wvlichtmast, unique_index=1)
    my_list_OTLAssets = [OTLObject for OTLObject in my_list_OTLObjects if not is_relation(OTLObject)]
    my_list_OTLRelations = [OTLObject for OTLObject in my_list_OTLObjects if is_relation(OTLObject)]

    with subtests.test(msg='Function returns a list'):
        # Gegeven een nieuwe asset, worden de juiste assets gegenereerd met de postenMapping + relaties + subassets.
        assert isinstance(my_list_OTLObjects, list)

    with subtests.test(msg='List elements are OTL Assets or OTL Relations'):
        assert (all(my_otlasset.is_instance_of(OTLAsset) for my_otlasset in my_list_OTLAssets),
                "Not all elements are instances of OTLAssets")
        assert (all(is_relation(my_otlrelation) for my_otlrelation in my_list_OTLRelations),
                "Not all elements are instances of OTLRelation")
        # number fo the OTLObjects is perfectly split in Assets and Relations
        assert (len(my_list_OTLObjects) == len(my_list_OTLAssets) + len(my_list_OTLRelations),
                "The number of created OTLObjects equals the number of OTLAssets and OTLRelations")

    with subtests.test(msg='All the OTL Assets have at least one value for attribute "toestand"'):
        # if toestand is missing (None), the length of the set is zero (0).
        # if toestand has a value, the length of the set should be 1 (all values identical)
        # or more than 1 (all the values are different)
        assert (len({ my_otlasset.toestand for my_otlasset in my_list_OTLAssets}) > 0)

@pytest.fixture
def factory_postenmapping_template_202411() -> PostAssetFactory():
    this_directory = Path(__file__).parent
    return PostAssetFactory(
        this_directory / 'PostenMapping_template_202411.db',
        directory=this_directory,
        mapping_name='PostenMapping_template_202411'
    )

@pytest.fixture
def base_asset_WVLichtmast() -> OTLAsset:
    asset = WVLichtmast()
    asset.fill_with_dummy_data()
    asset.naam = 'myDummyNaam'
    asset.bestekPostNummer = ['WVlichtmast_config1']
    asset.geometry = 'POINT Z (160000 160000 0)'
    return asset

@pytest.mark.parametrize("overwrite_original_attributes_by_mapping, expected_naam, expected_geometry", [
    (True, None, None),   # Expect the original name to be overwritten (set to None)
    (False, 'myDummyNaam', 'POINT Z (160000 160000 0)')   # Expect the original attribute "naam" and "geometry" to be retained
])
def test_create_asset_from_mapping_overwrite_or_keep_attributes(factory_postenmapping_template_202411, base_asset_WVLichtmast, overwrite_original_attributes_by_mapping, expected_naam, expected_geometry, subtests):
    my_list_OTLObjects = factory_postenmapping_template_202411.create_assets_from_mapping(
        base_asset=base_asset_WVLichtmast, unique_index=1,
        overwrite_original_attributes_by_mapping=overwrite_original_attributes_by_mapping)
    my_list_OTLAssets = [obj for obj in my_list_OTLObjects if obj.is_instance_of(OTLAsset)]
    base_asset_from_list = my_list_OTLAssets[0]  # First asset in the list is the base asset

    with subtests.test(msg=f'Attribute "naam" is {"overwritten" if overwrite_original_attributes_by_mapping else "preserved"}'):
        assert base_asset_from_list.naam == expected_naam

    with subtests.test(msg=f'Attribute "geometry" is {"overwritten" if overwrite_original_attributes_by_mapping else "preserved"}'):
        assert base_asset_from_list.geometry == expected_geometry

@pytest.mark.parametrize(
    "overwrite_original_attributes_by_mapping",
    [[True], [False]]
)
def test_create_assets_from_mapping_and_write_to_file_overwrite_original_attributes(subtests, overwrite_original_attributes_by_mapping):
    # arrange
    this_directory = Path(__file__).parent
    file_path = this_directory / 'output.xlsx'
    model_directory_path = Path(__file__).parent / 'TestModel'
    factory = PostAssetFactory()
    factory.mapping_dict = TestModelPostenMappingDict.mapping_dict

    instance = AllCasesTestClass()
    instance.bestekPostNummer = ['testclass_keep_attributes']
    instance.testIntegerField = -9
    instance.testStringField = 'myDummyInitialString'
    start_assets = [instance]

    # act
    factory.create_assets_from_mapping_and_write_to_file(start_assets=start_assets, output_path=file_path,
                                                         overwrite_original_attributes_by_mapping=overwrite_original_attributes_by_mapping,
                                                         append_all_attributes=True,
                                                         model_directory=model_directory_path)

    # assert
    with subtests.test(msg=f'Output file exists: {file_path}'):
        assert os.path.exists(file_path)

    # Temporary delete the Sheet "Keuzelijsten"
    wb = load_workbook(file_path)
    wb.remove(wb['Keuzelijsten'])
    wb.save(file_path)

    instance_generated = list(OtlmowConverter.from_file_to_objects(file_path, model_directory=model_directory_path))[0]

    # remove attributes 'bestekPostNummer' from both instances
    instance_dict = instance.create_dict_from_asset()
    instance_dict.pop('bestekPostNummer', None)
    instance_generated_dict = instance_generated.create_dict_from_asset()
    instance_generated_dict.pop('bestekPostNummer', None)

    # Add specific assertions based on expected behavior
    if overwrite_original_attributes_by_mapping:
        with subtests.test(msg=f'test parameter: overwrite_original_attributes_by_mapping={overwrite_original_attributes_by_mapping}'):
            assert instance_generated.testIntegerField == 9
            assert instance_generated.testStringField == 'myDummyString'
    else:
        with subtests.test(msg=f'test parameter: overwrite_original_attributes_by_mapping={overwrite_original_attributes_by_mapping}'):
            assert instance_generated.testIntegerField == -9
            assert instance_generated.testStringField == 'myDummyInitialString'


@pytest.mark.parametrize(
    "append_all_attributes",
    [[True], [False]]
)
def test_create_assets_from_mapping_and_write_to_file_append_all_attributes(subtests, append_all_attributes):
    # arrange
    this_directory = Path(__file__).parent
    file_path = this_directory / 'output.xlsx'
    model_directory_path = Path(__file__).parent / 'TestModel'
    factory = PostAssetFactory()
    factory.mapping_dict = TestModelPostenMappingDict.mapping_dict

    instance = AllCasesTestClass()
    instance.bestekPostNummer = ['testclass_keep_attributes']
    instance.testIntegerField = -9
    instance.testStringField = 'myDummyInitialString'
    start_assets = [instance]

    # act
    factory.create_assets_from_mapping_and_write_to_file(start_assets=start_assets, output_path=file_path,
                                                         overwrite_original_attributes_by_mapping=True,
                                                         append_all_attributes=append_all_attributes,
                                                         model_directory=model_directory_path)

    # assert
    with subtests.test(msg=f'Output file exists: {file_path}'):
        assert os.path.exists(file_path)

    # Temporary delete the Sheet "Keuzelijsten"
    wb = load_workbook(file_path)
    wb.remove(wb['Keuzelijsten'])
    wb.save(file_path)

    instance_generated = list(OtlmowConverter.from_file_to_objects(file_path, model_directory=model_directory_path))[0]

    # remove attributes 'bestekPostNummer' from both instances
    instance_dict = instance.create_dict_from_asset()
    instance_dict.pop('bestekPostNummer', None)
    instance_generated_dict = instance_generated.create_dict_from_asset()
    instance_generated_dict.pop('bestekPostNummer', None)

    # Add specific assertions based on expected behavior
    wb = load_workbook(file_path)
    nbr_columns_excel = wb['ond#AllCasesTestClass'].max_column
    if append_all_attributes:
        # Open Excel sheet and count the number of (empty) columns.
        # This should be >= the number of attributes
        with subtests.test(msg=f'test parameter: append_all_attributes={append_all_attributes}'):
            assert nbr_columns_excel >= len(instance_dict)
    else:
        with subtests.test(msg=f'test parameter: append_all_attributes={append_all_attributes}'):
            assert nbr_columns_excel == len(instance_dict)

